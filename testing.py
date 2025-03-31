from re import search
from email.message import Message
from time import process_time_ns



import telebot
import openpyxl
from openpyxl import load_workbook
from telebot import types




Nbooking = openpyxl.open("11_11.xlsx")
Nsheeting = Nbooking["ГРУППЫ"]


booking =openpyxl.open("11_11.xlsx")
sheeting = booking["ГРУППЫ"]



bot = telebot.TeleBot("XXXXXXXXXXXXXXX")




word = {}

@bot.message_handler(commands=['start'])
def start(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("Четная")
    btn2 = types.KeyboardButton("Нечетная")
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, text="Привет, {0.first_name}! Я тестовый бот расписания".format(message.from_user), reply_markup=markup)

@bot.message_handler(content_types=['text'])
def week_sending(message):
	if (message.text == "Четная"):


		bot.send_message(message.chat.id, text="Выбирай день")
		markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
		btn1 = types.KeyboardButton(".Понедельник.")
		btn2 = types.KeyboardButton(".Вторник.")
		btn3 = types.KeyboardButton(".Среда.")
		btn4 = types.KeyboardButton(".Четверг.")
		btn5 = types.KeyboardButton(".Пятница.")
		back = types.KeyboardButton("Назад")
		markup.add(btn1, btn2, btn3,btn4,btn5,back)
		bot.send_message(message.chat.id,text="-",reply_markup=markup)
	elif (message.text == "Нечетная"):
		bot.send_message(message.chat.id, text="Выбирай день")
		markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
		btn1 = types.KeyboardButton("Понедельник")
		btn2 = types.KeyboardButton("Вторник")
		btn3 = types.KeyboardButton("Среда")
		btn4 = types.KeyboardButton("Четверг")
		btn5 = types.KeyboardButton("Пятница")
		back = types.KeyboardButton("Назад")
		markup.add(btn1, btn2, btn3, btn4, btn5, back)
		bot.send_message(message.chat.id, text="-", reply_markup=markup)
	elif (message.text == "Понедельник"):
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, Nhandle_mondey)
	elif message.text == "Вторник":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, Nhandle_tuesdey)
	elif message.text == "Среда":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, Nhandle_WEDNESDEY)
	elif message.text == "Четверг":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, Nhandle_thursdey)
	elif message.text == "Пятница":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, Nhandle_fridey)

	elif (message.text == ".Понедельник."):
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, handle_mondey)
	elif message.text == ".Вторник.":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, handle_tuesdey)
	elif message.text == ".Среда.":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, handle_WEDNESDEY)
	elif message.text == ".Четверг.":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, handle_thursdey)
	elif message.text == ".Пятница.":
		bot.send_message(message.chat.id, text="Введите группу для поиска(В ФОРМАТЕ:216-ИС-23: ")
		bot.register_next_step_handler(message, handle_fridey)


	elif (message.text == "Назад"):
		markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
		button1 = types.KeyboardButton("Четная")
		button2 = types.KeyboardButton("Нечетная")
		markup.add(button1, button2)
		bot.send_message(message.chat.id, text="Вы вернулись в главное меню", reply_markup=markup)


#для ЧЕТНОЙ недели
def handle_mondey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(sheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break

		if found:
			break

	if found is False:
		if (message.text)  == "Назад":
		   week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, handle_mondey)






	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = sheeting[row1][0].value
		B = sheeting[row1][1].value
		C = sheeting[row1][2].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if sheeting[row1][0].value is None:
				print('')
			elif sheeting[row1][1].value is None:
				print('')
			elif sheeting[row1][2].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{sheeting[row1][0].value} пара  " +f"{sheeting[row1][1].value} " +f"{sheeting[row1][2].value}")
def handle_tuesdey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(sheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, handle_tuesdey)
	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = sheeting[row1][0].value
		D = sheeting[row1][3].value
		E = sheeting[row1][4].value
		list = [A,D,E]
		if None or '' in list:
			print('')
		else:
			if sheeting[row1][0].value is None:
				print('')
			elif sheeting[row1][3].value is None:
				print('')
			elif sheeting[row1][4].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{sheeting[row1][0].value} пара  " +f"{sheeting[row1][3].value} " +f"{sheeting[row1][4].value}")
def handle_WEDNESDEY(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(sheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, handle_WEDNESDEY)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = sheeting[row1][0].value
		F = sheeting[row1][5].value
		G = sheeting[row1][6].value
		list = [A,F,G]
		if None or '' in list:
			print('')
		else:
			if sheeting[row1][0].value is None:
				print('')
			elif sheeting[row1][5].value is None:
				print('')
			elif sheeting[row1][6].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{sheeting[row1][0].value} пара  " +f"{sheeting[row1][5].value} " +f"{sheeting[row1][6].value}")
def handle_thursdey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(sheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, handle_thursdey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = sheeting[row1][0].value
		B = sheeting[row1][7].value
		C = sheeting[row1][8].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if sheeting[row1][0].value is None:
				print('')
			elif sheeting[row1][7].value is None:
				print('')
			elif sheeting[row1][8].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{sheeting[row1][0].value} пара  " +f"{sheeting[row1][7].value} " +f"{sheeting[row1][8].value}")
def handle_fridey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(sheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, handle_fridey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = sheeting[row1][0].value
		B = sheeting[row1][9].value
		C = sheeting[row1][10].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if sheeting[row1][0].value is None:
				print('')
			elif sheeting[row1][9].value is None:
				print('')
			elif sheeting[row1][10].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{sheeting[row1][0].value} пара  " +f"{sheeting[row1][9].value} " +f"{sheeting[row1][10].value}")






#для НЕЕЕЕЕЕЕЕЧЕТНОЙ недели
def Nhandle_mondey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(Nsheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, Nhandle_mondey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = Nsheeting[row1][0].value
		B = Nsheeting[row1][1].value
		C = Nsheeting[row1][2].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if Nsheeting[row1][0].value is None:
				print('')
			elif Nsheeting[row1][1].value is None:
				print('')
			elif Nsheeting[row1][2].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{Nsheeting[row1][0].value} пара  " +f"{Nsheeting[row1][1].value} " +f"{Nsheeting[row1][2].value}")
def Nhandle_tuesdey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(Nsheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, Nhandle_tuesdey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = Nsheeting[row1][0].value
		D = Nsheeting[row1][3].value
		E = Nsheeting[row1][4].value
		list = [A,D,E]
		if None or '' in list:
			print('')
		else:
			if Nsheeting[row1][0].value is None:
				print('')
			elif Nsheeting[row1][3].value is None:
				print('')
			elif Nsheeting[row1][4].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{Nsheeting[row1][0].value} пара  " +f"{Nsheeting[row1][3].value} " +f"{Nsheeting[row1][4].value}")
def Nhandle_WEDNESDEY(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(Nsheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, Nhandle_WEDNESDEY)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = Nsheeting[row1][0].value
		F = Nsheeting[row1][5].value
		G = Nsheeting[row1][6].value
		list = [A,F,G]
		if None or '' in list:
			print('')
		else:
			if Nsheeting[row1][0].value is None:
				print('')
			elif Nsheeting[row1][5].value is None:
				print('')
			elif Nsheeting[row1][6].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{Nsheeting[row1][0].value} пара  " +f"{Nsheeting[row1][5].value} " +f"{Nsheeting[row1][6].value}")
def Nhandle_thursdey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(Nsheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, Nhandle_thursdey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = Nsheeting[row1][0].value
		B = Nsheeting[row1][7].value
		C = Nsheeting[row1][8].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if Nsheeting[row1][0].value is None:
				print('')
			elif Nsheeting[row1][7].value is None:
				print('')
			elif Nsheeting[row1][8].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{Nsheeting[row1][0].value} пара  " +f"{Nsheeting[row1][7].value} " +f"{Nsheeting[row1][8].value}")
def Nhandle_fridey(message):
	word[message.chat.id] = message.text
	print("10%")
	search_word = f"Группа - " + word[message.chat.id]
	print("20%")
	found = False
	for row_number, row in enumerate(Nsheeting.iter_rows(), start=1):
		for cell in row:
			if cell.value == search_word:  # Если значение ячейки совпадает с искомым словом
				print(search_word)
				print("50%")

				found = True
				break
		if found:
			break

	if found is False:
		if (message.text) == "Назад":
			week_sending(message)
		else:
			print(f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Слово '{search_word}' не найдено.")
			bot.send_message(message.chat.id, text=f"Попробуйте еще раз")
			bot.send_message(message.chat.id, text="Введите группу для поиска В ФОРМАТЕ 216-ИС-23 ")

			bot.register_next_step_handler(message, Nhandle_fridey)

	row_number2 = row_number + 6
	row_number1 = row_number + 29
	for row1 in range(row_number2, row_number1):

		A = Nsheeting[row1][0].value
		B = Nsheeting[row1][9].value
		C = Nsheeting[row1][10].value
		list = [A,B,C]
		if None or '' in list:
			print('')
		else:
			if Nsheeting[row1][0].value is None:
				print('')
			elif Nsheeting[row1][9].value is None:
				print('')
			elif Nsheeting[row1][10].value is None:
				print('')

			else:
				bot.send_message(message.chat.id,text=f"{Nsheeting[row1][0].value} пара  " +f"{Nsheeting[row1][9].value} " +f"{Nsheeting[row1][10].value}")

















if __name__ == "__main__":
	bot.polling(none_stop=True)
